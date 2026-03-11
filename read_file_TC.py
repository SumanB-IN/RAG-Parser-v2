import pandas as pd
import re
from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook


class ReadFileTC:
# Define the file path for your Excel file\
    def __init__(self):
        # self.path = path
        # self.sheet_data = None 
        pass
    
    def normalized_text(self, value: Any) -> str:
        """Return trimmed string form of a cell, or empty string when blank/NaN."""
        if pd.isna(value):
            return ""
        return str(value).strip()

    def is_natural_number(self, value: Any) -> bool:
        """Check whether a value is a natural number (positive integer, not Roman numeral text)."""
        if pd.isna(value):
            return False

        if isinstance(value, bool):
            return False

        if isinstance(value, int):
            return value > 0

        if isinstance(value, float):
            return value.is_integer() and value > 0

        text = str(value).strip().replace(",", "")
        if text == "":
            return False
        if not re.fullmatch(r"\d+", text):
            return False
        return int(text) > 0

    def sanitize_filename(self, excel_file_path:str, name: str) -> str: #Modified by Suman
        """Convert arbitrary equipment names into Windows-safe filenames."""
        file_name = Path(excel_file_path).stem
        cleaned = str(name).strip()
        cleaned = re.sub(r'[<>:"/|?*\x00-\x1F]', "_", cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
        cleaned = file_name + "$" + cleaned 
        return cleaned or "unnamed_equipment"

    def find_ser_no_location(self, raw_df: pd.DataFrame) -> tuple[int, int] | None:
        """Find the (row, col) location of 'Ser No', 'Ser', or 'S No' in rows 1..6 (0-based 0..5)."""
        max_rows = min(6, len(raw_df))
        for row_idx in range(max_rows):
            row = raw_df.iloc[row_idx]
            for col_idx, value in enumerate(row.tolist()):
                token = self.normalized_text(value).casefold()
                if token in {"ser no", "ser", "s no"}:
                    return row_idx, col_idx
        return None

    def get_frs_hierarchy_path(self, excel_path: Path) -> Path:
        """
        Return path from 'FRS' to the Excel file (inclusive), if present.

        Example:
        .../FRS/Nov/Fmn C Nov/APPX Nov 2025.xlsx -> FRS/Nov/Fmn C Nov/APPX Nov 2025.xlsx
        """
        parts = list(excel_path.parts)
        for idx, part in enumerate(parts):
            if str(part).casefold() == "frs":
                return Path(*parts[idx:])
        return Path(excel_path.name)
    
    def header_text_with_merges(self,
        ws: Any, raw_df: pd.DataFrame, row_idx: int, col_idx: int, backend: str
    ) -> str:
        """
        Return header text for a given cell, honoring true Excel merged ranges only.

        If pandas gives an empty value for a merged child cell, use the merged area's
        top-left value. Otherwise keep it empty.
        """
        if row_idx >= len(raw_df) or col_idx >= raw_df.shape[1]:
            return ""

        direct = self.normalized_text(raw_df.iat[row_idx, col_idx])
        if direct:
            return direct

        if backend == "openpyxl":
            row_1based = row_idx + 1
            col_1based = col_idx + 1
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row_1based <= merged_range.max_row
                    and merged_range.min_col <= col_1based <= merged_range.max_col
                ):
                    top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
                    return self.normalized_text(top_left)
        else:
            # xlrd merged cell tuples are (row_start, row_end_exclusive, col_start, col_end_exclusive)
            for rlo, rhi, clo, chi in ws.merged_cells:
                if rlo <= row_idx < rhi and clo <= col_idx < chi:
                    return self.normalized_text(ws.cell_value(rlo, clo))

        return ""
    
    def build_cell_names(self, raw_df: pd.DataFrame, ws: Any, header_row_idx: int, backend: str) -> list[str]:
        """Build cell names by concatenating labels from two adjacent header rows."""
        cell_names: list[str] = []
        seen: Dict[str, int] = {}

        for col_idx in range(raw_df.shape[1]):
            part1 = self.header_text_with_merges(ws, raw_df, header_row_idx, col_idx, backend)
            part2 = self.header_text_with_merges(ws, raw_df, header_row_idx + 1, col_idx, backend)
            parts = [part for part in (part1, part2) if part]

            if parts:
                base_name = "/".join(parts)
            else:
                base_name = f"column_{col_idx + 1}"

            count = seen.get(base_name, 0) + 1
            seen[base_name] = count
            final_name = base_name if count == 1 else f"{base_name}_{count}"
            cell_names.append(final_name)

        return cell_names
    
    def first_match(self, named_values: Dict[str, Any], predicates: list) -> Any:
        """Return first value whose normalized key matches any predicate."""
        for key, value in named_values.items():
            norm = self.canonicalize_name(key)
            if any(predicate(norm) for predicate in predicates):
                return value
        return 0
    
    def row_to_cell_entries(self, row: pd.Series, cell_names: list[str]) -> list[Dict[str, Any]]:
        """Convert row to list of {cell_name, cell_contents}, treating empty cells as 0."""
        entries: list[Dict[str, Any]] = []
        for col_idx, cell_name in enumerate(cell_names):
            value = row.iloc[col_idx] if col_idx < len(row) else None
            if pd.isna(value) or str(value).strip() == "":
                value = 0
            entries.append({"cell_name": cell_name, "cell_contents": value})
        return entries

    def row_to_named_values(self, row: pd.Series, cell_names: list[str]) -> Dict[str, Any]:
        """Map computed cell names to row values, treating empty cells as 0."""
        values: Dict[str, Any] = {}
        for col_idx, cell_name in enumerate(cell_names):
            value = row.iloc[col_idx] if col_idx < len(row) else None
            if pd.isna(value) or str(value).strip() == "":
                value = 0
            values[cell_name] = value
        return values

    def canonicalize_name(self, name: str) -> str:
        """Normalize a cell name for rule-based matching."""
        cleaned = re.sub(r"[^a-z0-9]+", " ", name.casefold()).strip()
        return re.sub(r"\s+", " ", cleaned)

    def build_flat_payload(self,
        sheet_name: str,
        ser_no_value: Any,
        equipment_name: str,
        frs_hierarchy: Path,
        named_values: Dict[str, Any],
        ) -> Dict[str, Any]:
        """Build flattened output JSON schema requested by user."""
        remarks_value = self.first_match(
            named_values,
            [
                lambda n: n.startswith("remarks "),
                lambda n: n == "remarks",
            ],
        )
        if remarks_value == 0:
            remarks_value = ""

        return {
            "sheet": sheet_name,
            "ser_no": int(float(str(ser_no_value).replace(",", ""))),
            "equipment_name": equipment_name,
            "source_file": str(frs_hierarchy),
            "auth_ue": self.first_match(named_values, [lambda n: "auth ue" in n]),
            "held_uh": self.first_match(named_values, [lambda n: "held uh" in n]),
            "mua": self.first_match(named_values, [lambda n: n.endswith(" mua"), lambda n: n == "mua"]),
            "oh": self.first_match(named_values, [lambda n: n.endswith(" oh"), lambda n: n == "oh"]),
            "r4": self.first_match(named_values, [lambda n: n.endswith(" r4"), lambda n: n == "r4"]),
            "total_nmc": self.first_match(
                named_values,
                [
                    lambda n: "dependency total" in n,
                    lambda n: "nmc total" in n,
                    lambda n: n == "total",
                ],
            ),
            "fmc": self.first_match(
                named_values,
                [
                    lambda n: "fmc total" in n,
                    lambda n: n == "fmc",
                ],
            ),
            "remarks": remarks_value,
        }

    def load_sheets_and_workbook(self, excel_path: Path) -> tuple[Dict[str, pd.DataFrame], Any, str]:
        """
        Load workbook data for both pandas rows and merged-cell metadata.

        Returns:
        - sheets: dict of sheet_name -> raw dataframe (header=None)
        - workbook: backend workbook object (openpyxl or xlrd)
        - backend: "openpyxl" or "xlrd"
        """
        suffix = excel_path.suffix.casefold()

        if suffix == ".xls":
            try:
                import xlrd  # type: ignore
            except ImportError as exc:
                raise ImportError(
                    "This file is .xls and requires xlrd. Install it with: pip install xlrd>=2.0.1"
                ) from exc

            sheets = pd.read_excel(excel_path, sheet_name=None, header=None, engine="xlrd")
            workbook = xlrd.open_workbook(str(excel_path))
            return sheets, workbook, "xlrd"

        sheets = pd.read_excel(excel_path, sheet_name=None, header=None, engine="openpyxl")
        workbook = load_workbook(excel_path, data_only=True)
        return sheets, workbook, "openpyxl"

