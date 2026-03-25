from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


class DataGenerator:
    def __init__(self, base_dir: str | Path | None = None):
        if base_dir is None:
            base_dir = Path(__file__).resolve().parent
        self.base_dir = Path(base_dir)

    def read_workorderlocal(self) -> pd.DataFrame:
        workbook_path = self.base_dir / "FRS_filtered" / "WorkOrderLocal.xlsx"

        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}. Expected file name: WorkOrderLocal.xlsx")

        workbook = load_workbook(workbook_path, data_only=True)
        worksheet = workbook.active

        for merged_range in list(worksheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = worksheet.cell(row=min_row, column=min_col).value
            worksheet.unmerge_cells(str(merged_range))

            for row_index in range(min_row, max_row + 1):
                for col_index in range(min_col, max_col + 1):
                    worksheet.cell(row=row_index, column=col_index).value = top_left_value

        values = list(worksheet.values)
        if len(values) < 2:
            return pd.DataFrame()

        header_row_2 = values[1]
        selected_indices = []
        headers = []

        for column_index, header_value in enumerate(header_row_2):
            if header_value is None:
                continue

            header_text = str(header_value).strip()
            header_lower = header_text.lower()
            if not header_text or header_lower.startswith("unnamed"):
                continue

            selected_indices.append(column_index)
            headers.append(header_text)

        rows = values[2:]
        filtered_rows = []
        for row in rows:
            filtered_rows.append([
                row[column_index] if column_index < len(row) else None
                for column_index in selected_indices
            ])

        dataframe = pd.DataFrame(filtered_rows, columns=headers)
        if not dataframe.empty:
            dataframe = dataframe.iloc[1:].reset_index(drop=True)

        first_three_columns = list(dataframe.columns[:3])
        if first_three_columns:
            dataframe = dataframe.dropna(subset=first_three_columns)

        dataframe = dataframe.ffill(axis=0)
        return dataframe

    def extrapolate_dataframe(self, dataframe: pd.DataFrame, target_rows: int = 100) -> pd.DataFrame:
        if target_rows <= 0:
            return dataframe.iloc[0:0].copy()

        if dataframe.empty:
            return dataframe.copy()

        base_dataframe = dataframe.reset_index(drop=True).copy()
        current_rows = len(base_dataframe)

        if current_rows >= target_rows:
            return base_dataframe.iloc[:target_rows].copy()

        numeric_columns = list(base_dataframe.select_dtypes(include="number").columns)
        last_row = base_dataframe.iloc[-1].copy()

        numeric_steps: dict[str, float] = {}
        for column in numeric_columns:
            column_series = base_dataframe[column].dropna()
            if len(column_series) >= 2:
                numeric_steps[column] = column_series.iloc[-1] - column_series.iloc[-2]
            else:
                numeric_steps[column] = 0

        rows_to_add = target_rows - current_rows
        generated_rows = []
        previous_row = last_row.copy()

        for _ in range(rows_to_add):
            new_row = previous_row.copy()
            for column in numeric_columns:
                step = numeric_steps.get(column, 0)
                previous_value = previous_row[column]
                if pd.isna(previous_value):
                    previous_value = base_dataframe[column].dropna().iloc[-1] if not base_dataframe[column].dropna().empty else 0
                new_row[column] = previous_value + step

            generated_rows.append(new_row)
            previous_row = new_row

        generated_dataframe = pd.DataFrame(generated_rows, columns=base_dataframe.columns)
        extrapolated_dataframe = pd.concat([base_dataframe, generated_dataframe], ignore_index=True)
        extrapolated_dataframe = extrapolated_dataframe.ffill(axis=0)
        return extrapolated_dataframe

    def save_dataframe_to_csv(self, dataframe: pd.DataFrame, output_file: str | Path = "workorderlocal_extrapolated.csv") -> Path:
        output_path = self.base_dir / output_file
        dataframe.to_csv(output_path, index=False)
        return output_path


if __name__ == "__main__":
    generator = DataGenerator()
    df = generator.read_workorderlocal()
    df_100 = generator.extrapolate_dataframe(df, target_rows=100)
    saved_path = generator.save_dataframe_to_csv(df_100)
    print(df_100.head())
    print(f"Rows: {len(df_100)}")
    print(f"Saved CSV: {saved_path}")